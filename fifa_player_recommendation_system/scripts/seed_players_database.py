import os
import psycopg2

from fifa_player_recommendation_system.modules import db
from openpyxl import load_workbook

os.chdir('..')

players_wb = load_workbook(os.getcwd() + '/data/players.xlsx')
players_ws = players_wb['Players']

teams_wb = load_workbook(os.getcwd() + '/data/teams.xlsx')
teams_ws = teams_wb['Teams']

'''
Go through the teams workbook first in order to create a dictionary where
the key is the name of the team and the value is the name of the league 
where that team players
'''
teams_dict = {
    'Free agent': ''
}
num_teams = teams_ws.max_row

for i in range(2, num_teams + 1):
    teams_dict[teams_ws['A{}'.format(i)].value] = teams_ws['B{}'.format(i)].value

conn = db.connect()
cursor = conn.cursor()

num_players = players_ws.max_row
for i in range (2, num_players + 1):
    player_id = None

    # separating each needed column into its own variable for easier readability in the insert_stmt variable

    # players variables
    short_name = players_ws['A{}'.format(i)].value
    full_name = players_ws['B{}'.format(i)].value
    overall = players_ws['C{}'.format(i)].value
    potential = players_ws['D{}'.format(i)].value
    value = players_ws['E{}'.format(i)].value
    position = players_ws['G{}'.format(i)].value
    nationality = players_ws['H{}'.format(i)].value
    picture_url = players_ws['I{}'.format(i)].value
    age = players_ws['J{}'.format(i)].value
    height = players_ws['K{}'.format(i)].value
    weight = players_ws['L{}'.format(i)].value
    club = players_ws['O{}'.format(i)].value
    league = teams_dict[club]
    wage = players_ws['P{}'.format(i)].value
    release_clause = players_ws['Q{}'.format(i)].value
    preferred_foot = players_ws['W{}'.format(i)].value

    # player_stats variables
    game_version = 23
    weak_foot = players_ws['X{}'.format(i)].value
    skill_moves = players_ws['Y{}'.format(i)].value
    attacking_work_rate = players_ws['AE{}'.format(i)].value
    defensive_work_rate = players_ws['AF{}'.format(i)].value
    crossing = players_ws['AM{}'.format(i)].value
    finishing = players_ws['AN{}'.format(i)].value
    heading_accuracy = players_ws['AO{}'.format(i)].value
    short_passing = players_ws['AP{}'.format(i)].value
    volleys = players_ws['AQ{}'.format(i)].value
    dribbling = players_ws['AR{}'.format(i)].value
    curve = players_ws['AS{}'.format(i)].value
    fk_accuracy = players_ws['AT{}'.format(i)].value
    long_passing = players_ws['AU{}'.format(i)].value
    ball_control = players_ws['AV{}'.format(i)].value
    acceleration = players_ws['AW{}'.format(i)].value
    sprint_speed = players_ws['AX{}'.format(i)].value
    agility = players_ws['AY{}'.format(i)].value
    reactions = players_ws['AZ{}'.format(i)].value
    balance = players_ws['BA{}'.format(i)].value
    shot_power = players_ws['BB{}'.format(i)].value
    jumping = players_ws['BC{}'.format(i)].value
    stamina = players_ws['BD{}'.format(i)].value
    strength = players_ws['BE{}'.format(i)].value
    long_shots = players_ws['BF{}'.format(i)].value
    aggression = players_ws['BG{}'.format(i)].value
    interceptions = players_ws['BH{}'.format(i)].value
    positioning = players_ws['BI{}'.format(i)].value
    vision = players_ws['BJ{}'.format(i)].value
    penalties = players_ws['BK{}'.format(i)].value
    composure = players_ws['BL{}'.format(i)].value
    marking = players_ws['BM{}'.format(i)].value
    standing_tackle = players_ws['BN{}'.format(i)].value
    sliding_tackle = players_ws['BO{}'.format(i)].value
    gk_diving = players_ws['BP{}'.format(i)].value
    gk_handling = players_ws['BQ{}'.format(i)].value
    gk_kicking = players_ws['BR{}'.format(i)].value
    gk_positioning = players_ws['BS{}'.format(i)].value
    gk_reflexes = players_ws['BT{}'.format(i)].value
    pace_total = players_ws['AG{}'.format(i)].value
    shooting_total = players_ws['AH{}'.format(i)].value
    passing_total = players_ws['AI{}'.format(i)].value
    dribbling_total = players_ws['AJ{}'.format(i)].value
    defending_total = players_ws['AK{}'.format(i)].value
    physicality_total = players_ws['AL{}'.format(i)].value
    st_rating = players_ws['BU{}'.format(i)].value
    lw_rating = players_ws['BV{}'.format(i)].value
    lf_rating = players_ws['BW{}'.format(i)].value
    cf_rating = players_ws['BX{}'.format(i)].value
    rf_rating = players_ws['BY{}'.format(i)].value
    rw_rating = players_ws['BZ{}'.format(i)].value
    cam_rating = players_ws['CA{}'.format(i)].value
    lm_rating = players_ws['CB{}'.format(i)].value
    cm_rating = players_ws['CC{}'.format(i)].value
    rm_rating = players_ws['CD{}'.format(i)].value
    lwb_rating = players_ws['CE{}'.format(i)].value
    cdm_rating = players_ws['CF{}'.format(i)].value
    rwb_rating = players_ws['CG{}'.format(i)].value
    lb_rating = players_ws['CH{}'.format(i)].value
    cb_rating = players_ws['CI{}'.format(i)].value
    rb_rating = players_ws['CJ{}'.format(i)].value
    gk_rating = players_ws['CK{}'.format(i)].value

    insert_player_stmt = 'INSERT INTO players(short_name, full_name, overall, potential,' \
        ' value, position, nationality, picture_url, age, height, weight, club, league,' \
        ' wage, release_clause, preferred_foot) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s,' \
        ' %s, %s, %s, %s, %s, %s, %s) RETURNING id'

    player = (short_name, full_name, overall, potential, value, position, nationality, \
        picture_url, age, height, weight, club, league, wage, release_clause, preferred_foot)

    try:
        cursor.execute(insert_player_stmt, player)
        player_id = cursor.fetchone()[0]
        conn.commit()
    except (Exception, psycopg2.DatabaseError) as error:
        print('Error occurred while executing the insert players statement')
        print(error)
    
    insert_player_stats_stmt = 'INSERT INTO player_stats(player_id, game_version, weak_foot,' \
        ' skill_moves, attacking_work_rate, defensive_work_rate, crossing, finishing, heading_accuracy,' \
        ' short_passing, volleys, dribbling, curve, freekick_accuracy, long_passing, ball_control,' \
        ' acceleration, sprint_speed, agility, reactions, balance, shot_power, jumping, stamina,' \
        ' strength, long_shots, aggression, interceptions, positioning, vision, penalties, composure,' \
        ' marking, standing_tackle, sliding_tackle, gk_diving, gk_handling, gk_kicking, gk_positioning,' \
        ' gk_reflexes, pace_total, shooting_total, passing_total, dribbling_total, defending_total,' \
        ' physicality_total, st_rating, lw_rating, lf_rating, cf_rating, rf_rating, rw_rating, cam_rating,' \
        ' lm_rating, cm_rating, rm_rating, lwb_rating, cdm_rating, rwb_rating, lb_rating, cb_rating,' \
        ' rb_rating, gk_rating) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,' \
        ' %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,' \
        ' %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)'

    player_stats = (player_id, game_version, weak_foot, skill_moves, attacking_work_rate, defensive_work_rate, \
        crossing, finishing, heading_accuracy, short_passing, volleys, dribbling, curve, fk_accuracy, long_passing, \
        ball_control, acceleration, sprint_speed, agility, reactions, balance, shot_power, jumping, stamina, \
        strength, long_shots, aggression, interceptions, positioning, vision, penalties, composure, \
        marking, standing_tackle, sliding_tackle, gk_diving, gk_handling, gk_kicking, gk_positioning, \
        gk_reflexes, pace_total, shooting_total, passing_total, dribbling_total, defending_total, \
        physicality_total, st_rating, lw_rating, lf_rating, cf_rating, rf_rating, rw_rating, cam_rating, \
        lm_rating, cm_rating, rm_rating, lwb_rating, cdm_rating, rwb_rating, lb_rating, cb_rating, \
        rb_rating, gk_rating)
    
    try:
        cursor.execute(insert_player_stats_stmt, player_stats)
        conn.commit()
    except (Exception, psycopg2.DatabaseError) as error:
        print('Error occurred while executing the insert player stats statement')
        print(error)

cursor.close()
conn.close()