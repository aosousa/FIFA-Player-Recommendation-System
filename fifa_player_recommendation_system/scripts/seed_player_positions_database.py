import os
import psycopg2

from fifa_player_recommendation_system.modules import db
from openpyxl import load_workbook

os.chdir('..')

players_wb = load_workbook(os.getcwd() + '/data/players.xlsx')
players_ws = players_wb['Players']

conn = db.connect()
cursor = conn.cursor()

num_players = players_ws.max_row
for i in range(2, num_players + 1):
    player_positions = players_ws['F{}'.format(i)].value.split(',')

    for position in player_positions:
        insert_player_position_stmt = 'INSERT INTO player_positions(player_id, position) VALUES (%s, %s)'

        player_position = (i-1, position)

        try:
            cursor.execute(insert_player_position_stmt, player_position)
            conn.commit()
        except (Exception, psycopg2.DatabaseError) as error:
            print('Error occurred while executing the insert player positions statement')
            print(error)

cursor.close()
conn.close()