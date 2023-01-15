import os
import psycopg2

from fifa_player_recommendation_system.modules import db
from openpyxl import load_workbook

os.chdir('..')

teams_wb = load_workbook(os.getcwd() + '/data/teams.xlsx')
teams_ws = teams_wb['Teams']

leagues = [league[0].value for league in teams_ws['B{}:B{}'.format(2, teams_ws.max_row)]]
unique_leagues = list(dict.fromkeys(leagues))

conn = db.connect()
insert_stmt = 'INSERT INTO leagues(name) VALUES'

for league in unique_leagues:
    insert_stmt += "('{}'), ".format(league)

insert_stmt = insert_stmt[0:len(insert_stmt) - 2] + ';'

try:
    cursor = conn.cursor()
    cursor.execute(insert_stmt)
    conn.commit()
    cursor.close()
except (Exception, psycopg2.DatabaseError) as error:
    print('Error occurred while executing the insert statement')
    print(error)
finally:
    if conn is not None:
        conn.close()