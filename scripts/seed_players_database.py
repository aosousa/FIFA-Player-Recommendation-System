import os

from openpyxl import load_workbook

os.chdir('..')

players_wb = load_workbook(os.getcwd() + '/data/players.xlsx')
players_ws = players_wb['Players']

teams_wb = load_workbook(os.getcwd() + '/data/teams.xlsx')
teams_ws = teams_wb['Teams']

'''
Go through the teams workbook first in order to create a dictionary where
the key is the name of the team and the value is the name of the league 
where that team players
'''
teams_dict = {}
num_teams = teams_ws.max_row

for i in range(2, num_teams):
    teams_dict[teams_ws['A{}'.format(i)].value] = teams_ws['B{}'.format(i)].value